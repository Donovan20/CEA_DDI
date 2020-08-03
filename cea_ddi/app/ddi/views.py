from django.shortcuts import render
from django.shortcuts import redirect
from django.db.models import Count
from django.http import HttpResponse
from django.contrib.auth.hashers import make_password
from django.http import HttpResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required

from wsgiref.util import FileWrapper

from django.contrib.auth.models import User
from cea_ddi.settings import BASE_DIR
from app.ddi.forms import UserForm
from app.ddi.models import Expediente
from app.ddi.forms import ExpForm
from app.ddi.models import Desarrolladora
from app.ddi.forms import DesaForm
from app.ddi.models import Categorias
from app.ddi.forms import CateForm
from app.ddi.models import SubCategorias
from app.ddi.forms import SubCateForm
from app.ddi.models import Proyectos
from app.ddi.forms import ProyeForm
from app.ddi.models import Ingresos
from app.ddi.forms import IngresoForm
from app.ddi.forms import RevisarForm
from app.ddi.models import Files
from app.ddi.models import Estados
from app.ddi.forms import NotaForm

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import matplotlib
import matplotlib.pyplot as plt
import numpy as np


import os
from datetime import timedelta
from datetime import datetime
from re import compile
from json import dumps


@login_required
def index(request):
    p = Proyectos.objects.filter(responsable=request.user).count()
    p2 = Proyectos.objects.filter(revisador=request.user).count
    return render(request, 'dashboard.html', {'proyectos': p, 'exp': p2})


@login_required
def lista_usuarios(request):
    usuarios = User.objects.all()
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            u = form.save(commit=False)
            if User.objects.filter(username=u.email):
                pass
            else:
                s = slice(3)
                p = u.first_name[s] + u.last_name[s] + ".@"
                u.password = make_password(p)
                u.username = u.email
                u.save()
                return redirect('/usuarios/')
    else:
        form = UserForm()
    return render(request, 'users.html', {'usuarios': usuarios, 'form': form})


@login_required
def editar_usuario(request, pk):
    u = User.objects.get(pk=pk)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=u)
        if form.is_valid():
            us = form.save(commit=False)
            if User.objects.filter(username=us.email):
                pass
            else:
                us.save()
                return redirect('/usuarios/')
    else:
        form = UserForm(instance=u)

    return render(request, 'editUser.html', {'form': form})


@login_required
def lista_expedientes(request):
    expedientes = Expediente.objects.all()
    if request.method == 'POST':
        form = ExpForm(request.POST)
        print(request.POST)
        if form.is_valid():
            exp = form.save(commit=False)
            folio = compile(
                r'^[A-z][A-z]\-[0-9][0-9][0-9]\-[0-9][0-9]\-[A-z]$')
            if folio.match(exp.numero):
                if Expediente.objects.filter(numero=exp.numero):
                    pass
                else:
                    exp.save()
                    return redirect('/expedientes/')
            else:
                pass
    else:
        form = ExpForm()
    return render(request, 'expedientes.html', {'expedientes': expedientes, 'form': form})


@login_required
def editar_expediente(request, pk):

    e = Expediente.objects.get(pk=pk)
    if request.method == 'POST':
        form = ExpForm(request.POST, instance=e)
        if form.is_valid():
            exp = form.save(commit=False)
            folio = compile(
                r'^[A-z][A-z]\-[0-9][0-9][0-9]\-[0-9][0-9]\-[A-z]$')
            if folio.match(exp.numero):
                if Expediente.objects.filter(numero=exp.numero):
                    pass
                else:
                    exp.save()
                    return redirect('/expedientes/')
    else:
        form = ExpForm(instance=e)
        print(form)
    return render(request, 'editExpediente.html', {'form': form})


@login_required
def eliminar(request):
    url = ""
    if request.POST['bandera'] == 'e':
        exp = Expediente.objects.get(pk=request.POST['pk'])
        exp.delete()
        url = '/expedientes/'

    elif request.POST['bandera'] == 'd':
        desa = Desarrolladora.objects.get(pk=request.POST['pk'])
        desa.delete()
        url = "/desarrolladoras/"

    elif request.POST['bandera'] == 'c':
        cate = Categorias.objects.get(pk=request.POST['pk'])
        cate.delete()
        url = "/categorias/"

    elif request.POST['bandera'] == 's':
        sub = SubCategorias.objects.get(pk=request.POST['pk'])
        sub.delete()
        url = "/subcategorias/"

    return HttpResponse(url)


@login_required
def lista_desarrolladoras(request):
    desarrolladoras = Desarrolladora.objects.all()
    if request.method == 'POST':
        form = DesaForm(request.POST)
        if form.is_valid():
            des = form.save(commit=False)
            if Desarrolladora.objects.filter(nombre=des.nombre):
                pass
            else:
                des.save()
                return redirect("/desarrolladoras/")
    else:
        form = DesaForm()
    return render(request, 'desarrolladora.html', {'desarrolladoras': desarrolladoras, 'form': form})


@login_required
def editar_desarrolladora(request, pk):
    d = Desarrolladora.objects.get(pk=pk)
    if request.method == 'POST':
        form = DesaForm(request.POST, instance=d)
        if form.is_valid():
            desa = form.save(commit=False)
            if Desarrolladora.objects.filter(nombre=desa.nombre):
                pass
            else:
                desa.save()
                return redirect('/desarrolladoras/')
    else:
        form = DesaForm(instance=d)
        print(form)
    return render(request, 'editDesarrolladora.html', {'form': form})


@login_required
def lista_categorias(request):
    categorias = Categorias.objects.all()
    if request.method == 'POST':
        form = CateForm(request.POST)
        if form.is_valid():
            cate = form.save(commit=False)
            if Categorias.objects.filter(nombre=cate.nombre):
                pass
            else:
                cate.save()
                return redirect("/categorias/")
    else:
        form = CateForm()
    return render(request, 'categorias.html', {'categorias': categorias, 'form': form})


@login_required
def editar_categoria(request, pk):
    c = Categorias.objects.get(pk=pk)
    if request.method == 'POST':
        form = CateForm(request.POST, instance=c)
        if form.is_valid():
            cate = form.save(commit=False)
            if Categorias.objects.filter(nombre=cate.nombre):
                pass
            else:
                cate.save()
                return redirect('/categorias/')
    else:
        form = CateForm(instance=c)

    return render(request, 'editCategoria.html', {'form': form})


@login_required
def lista_subcategorias(request):
    subcategorias = SubCategorias.objects.all()
    if request.method == 'POST':
        form = SubCateForm(request.POST)
        if form.is_valid():
            sub = form.save(commit=False)
            if SubCategorias.objects.filter(nombre=sub.nombre):
                pass
            else:
                sub.save()
                return redirect('/subcategorias/')
    else:
        form = SubCateForm()
    return render(request, 'subcategorias.html', {'subcategorias': subcategorias, 'form': form})


@login_required
def editar_subcategoria(request, pk):
    subcategoria = SubCategorias.objects.get(pk=pk)
    if request.method == 'POST':
        form = SubCateForm(request.POST, instance=subcategoria)
        if form.is_valid():
            sub = form.save(commit=False)
            if SubCategorias.objects.filter(nombre=sub.nombre):
                pass
            else:
                sub.save()
                return redirect('/subcategorias/')
    else:
        form = SubCateForm(instance=subcategoria)

    return render(request, 'editSubcategoria.html', {'form': form})


@login_required
def lista_proyectos(request):
    proyectos = Proyectos.objects.all()
    print(proyectos)
    if request.method == 'POST':
        form = ProyeForm(request.POST)
        if form.is_valid():
            proy = form.save(commit=False)
            if Proyectos.objects.filter(expediente=proy.expediente):
                pass
            else:
                proy.ingreso = 0
                proy.save()
                return redirect('/proyectos/')
    else:
        form = ProyeForm()
    return render(request, 'proyectos.html', {'proyectos': proyectos, 'form': form})


@login_required
def lista_ingresos(request, pk):
    ingresos = Ingresos.objects.filter(proyecto__pk=pk)
    return render(request, "ingresos.html", {'ingresos': ingresos})


@login_required
def revisar_ingreso(request, pk):
    ingreso = Ingresos.objects.get(pk=pk)
    if request.method == 'POST':
        form = RevisarForm(request.POST, request.FILES, instance=ingreso)
        if form.is_valid():
            i = form.save(commit=False)
            s = Estados.objects.get(pk=i.status.pk)
            p = Proyectos.objects.get(pk=i.proyecto.pk)
            path = "planos\\"+p.expediente.numero+"\\"+i.folio+"\\"
            i.status = s
            i.dias = int(abs((i.fecha_respuesta - i.fecha_ingreso).days))
            i.save()
            for file in request.FILES.getlist('files'):
                f = Files(file=file, ingreso=ingreso, tipo="Respuesta")
                f.file.storage.location = path
                f.save()
            return redirect('/proyectos/detalles/'+str(p.pk)+"/")
    else:
        form = RevisarForm(instance=ingreso)

    return render(request, 'revisar_ingreso.html', {'form': form})


@ login_required
def editar_proyecto(request, pk):
    p = Proyectos.objects.get(pk=pk)
    if request.method == 'POST':
        form = ProyeForm(request.POST, instance=p)
        if form.is_valid():
            print("paso")
            proye = form.save(commit=False)
            proye.save()
            return redirect('/proyectos/')
    else:
        form = ProyeForm(instance=p)

    return render(request, 'editProye.html', {'form': form})


@ login_required
def usuario_proyectos(request):
    proyectos = Proyectos.objects.filter(responsable=request.user)
    return render(request, 'mis_proyectos.html', {'proyectos': proyectos})


@ login_required
def usuario_ingresos(request, pk):
    ingresos = Ingresos.objects.filter(proyecto__pk=pk)
    if request.method == 'POST':
        form = IngresoForm(request.POST, request.FILES)
        if form.is_valid():
            i = form.save(commit=False)
            if Ingresos.objects.filter(folio=i.folio):
                pass
            else:
                s = Estados.objects.get(status='R')
                p = Proyectos.objects.get(pk=pk)
                path = "planos\\"+p.expediente.numero+"\\"
                if os.path.exists(os.path.join(BASE_DIR, path)):
                    path = path + i.folio+"\\"
                    if os.path.exists(os.path.join(BASE_DIR, path)):
                        pass
                    else:
                        os.mkdir(os.path.join(BASE_DIR, path))
                else:
                    os.mkdir(os.path.join(BASE_DIR, path))
                    path = path + i.folio+"\\"
                    if os.path.exists(os.path.join(BASE_DIR, path)):
                        pass
                    else:
                        os.mkdir(os.path.join(BASE_DIR, path))
                i.status = s
                i.proyecto = p
                d = timedelta(days=21)
                i.fecha_programada = i.fecha_ingreso + d
                i.ingreso = p.ingreso + 1
                i.save()
                for file in request.FILES.getlist('files'):
                    f = Files(file=file, ingreso=i, tipo="Ingreso")
                    f.file.storage.location = os.path.join(BASE_DIR, path)
                    f.save()
                p.ingreso += 1
                p.save()
                return redirect('/user/proyectos/ingresos/'+pk+"/")
    else:
        form = IngresoForm()

    return render(request, "mis_ingresos.html", {'ingresos': ingresos, 'form': form})


@ login_required
def ver_archivos(request, pk):
    """
    Send a file through Django without loading the whole file into
    memory at once. The FileWrapper will turn the file object into an
    iterator for chunks of 8KB.
    """
    ingreso = Ingresos.objects.get(pk=pk)
    fs = Files.objects.filter(ingreso__pk=pk).filter(tipo="Ingreso")
    files = []
    for f in fs:
        aux = {}
        path = "http://127.0.0.1:8000/planos/"+ingreso.proyecto.expediente.numero + \
            "/"+ingreso.folio+"/"+str(f.file)
        aux['path'] = path
        aux['name'] = str(f.file)
        files.append(aux)
    print(files)
    return render(request, "archivos.html", {'files': files})


@ login_required
def ver_oficios(request, pk):
    """
    Send a file through Django without loading the whole file into
    memory at once. The FileWrapper will turn the file object into an
    iterator for chunks of 8KB.
    """
    ingreso = Ingresos.objects.get(pk=pk)
    fs = Files.objects.filter(ingreso__pk=pk).filter(tipo="Respuesta")
    files = []
    for f in fs:
        aux = {}
        path = "http://127.0.0.1:8000/planos/"+ingreso.proyecto.expediente.numero + \
            "/"+ingreso.folio+"/"+str(f.file)
        aux['path'] = path
        aux['name'] = str(f.file)
        files.append(aux)
    print(files)
    return render(request, "archivos.html", {'files': files})


@ login_required
def ver_reportes(request):
    return render(request, "reports.html")


@ login_required
def reporte_nota_por_expediente(request):
    if request.method == "POST":
        fecha_I = request.POST['inicio']
        fecha_F = request.POST['fin']
        expediente = request.POST['expediente']
        categoria = request.POST['categoria']
        ingresos = Ingresos.objects.filter(
            Q(fecha_ingreso__gt=fecha_I) & Q(fecha_ingreso__lt=fecha_F)).filter(proyecto__tipo=categoria).filter(proyecto__expediente=expediente)
        expediente = Expediente.objects.get(pk=expediente).numero
        centrado = Alignment(horizontal='center')
        negritas = Font(bold=True)
        borde = Border(left=Side(border_style='thin',
                                 color='FF000000'),
                       right=Side(border_style='thin',
                                  color='FF000000'),
                       top=Side(border_style='thin',
                                color='FF000000'),
                       bottom=Side(border_style='thin',
                                   color='FF000000'))
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE INGRESOS DEL EXPEDIENTE: ' + \
            expediente
        celda = ws['A1']
        celda.alignment = centrado
        celda.font = negritas
        celda.border = borde
        ws.merge_cells('A1:G1')
        ws['A2'] = 'Expediente'
        celda = ws['A2']
        celda.alignment = centrado
        celda.font = negritas
        celda.border = borde
        ws['B2'] = 'Nombre del Proyecto'
        celda = ws['B2']
        celda.alignment = centrado
        celda.font = negritas
        celda.border = borde
        ws['C2'] = 'Estatus'
        celda = ws['C2']
        celda.alignment = centrado
        celda.font = negritas
        celda.border = borde
        ws['D2'] = 'Fecha de Ingreso'
        celda = ws['D2']
        celda.alignment = centrado
        celda.font = negritas
        celda.border = borde
        ws['E2'] = 'Fecha de respuesta'
        celda = ws['E2']
        celda.alignment = centrado
        celda.font = negritas
        celda.border = borde
        ws['F2'] = 'Oficio de respuesta'
        celda = ws['F2']
        celda.alignment = centrado
        celda.font = negritas
        celda.border = borde
        ws['G2'] = 'Observaciones'
        celda = ws['G2']
        celda.alignment = centrado
        celda.font = negritas
        celda.border = borde

        aux = 3

        for i in ingresos:
            ws.cell(row=aux, column=1).value = i.proyecto.expediente.numero
            ws.cell(row=aux, column=1).alignment = centrado
            ws.cell(row=aux, column=1).border = borde
            ws.cell(row=aux, column=2).value = i.proyecto.nombre
            ws.cell(row=aux, column=2).alignment = centrado
            ws.cell(row=aux, column=2).border = borde
            ws.cell(row=aux, column=3).value = i.status.nombre
            ws.cell(row=aux, column=3).border = borde
            ws.cell(row=aux, column=4).value = i.fecha_ingreso
            ws.cell(row=aux, column=4).alignment = centrado
            ws.cell(row=aux, column=4).border = borde
            ws.cell(row=aux, column=5).value = i.fecha_respuesta
            ws.cell(row=aux, column=5).border = borde
            ws.cell(row=aux, column=5).alignment = centrado
            ws.cell(row=aux, column=6).value = i.oficio
            ws.cell(row=aux, column=6).alignment = centrado
            ws.cell(row=aux, column=6).border = borde
            ws.cell(row=aux, column=7).value = i.observaciones
            ws.cell(row=aux, column=7).alignment = centrado
            ws.cell(row=aux, column=7).border = borde
            aux += 1

        nombre_archivo = "Nota_informativa_del_expediente_" + expediente + ".xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    else:
        form = NotaForm()
        print(form)
    return render(request, 'reporte_nota.html', {'form': form})


@ login_required
def reporte_ingresos_por_dia(request):
    hoy = datetime.now()
    ingresos = Ingresos.objects.filter(fecha_ingreso=hoy)
    centrado = Alignment(horizontal='center')
    negritas = Font(bold=True)
    borde = Border(left=Side(border_style='thin',
                             color='FF000000'),
                   right=Side(border_style='thin',
                              color='FF000000'),
                   top=Side(border_style='thin',
                            color='FF000000'),
                   bottom=Side(border_style='thin',
                               color='FF000000'))
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'REPORTE DE INGRESOS DEL DIA DE HOY: ' + \
        hoy.strftime("%d de %B del %Y")
    celda = ws['A1']
    celda.alignment = centrado
    celda.font = negritas
    celda.border = borde
    ws.merge_cells('A1:H1')
    ws['A2'] = 'Folio'
    celda = ws['A2']
    celda.alignment = centrado
    celda.font = negritas
    celda.border = borde
    ws['B2'] = 'Expediente'
    celda = ws['B2']
    celda.alignment = centrado
    celda.font = negritas
    celda.border = borde
    ws['C2'] = 'Nombre del Proyecto'
    celda = ws['C2']
    celda.alignment = centrado
    celda.font = negritas
    celda.border = borde
    ws['D2'] = 'Fecha de Ingreso'
    celda = ws['D2']
    celda.alignment = centrado
    celda.font = negritas
    celda.border = borde
    ws['E2'] = 'Desarrolladora'
    celda = ws['E2']
    celda.alignment = centrado
    celda.font = negritas
    celda.border = borde
    ws['F2'] = 'Representante legal'
    celda = ws['F2']
    celda.alignment = centrado
    celda.font = negritas
    celda.border = borde
    ws['G2'] = 'Proyectista CEA'
    celda = ws['G2']
    celda.alignment = centrado
    celda.font = negritas
    celda.border = borde
    ws['H2'] = 'Firma Proyectista'
    celda = ws['H2']
    celda.alignment = centrado
    celda.font = negritas
    celda.border = borde

    aux = 3

    for i in ingresos:
        ws.cell(row=aux, column=1).value = i.folio
        ws.cell(row=aux, column=1).alignment = centrado
        ws.cell(row=aux, column=1).border = borde
        ws.cell(row=aux, column=2).value = i.proyecto.expediente.numero
        ws.cell(row=aux, column=2).alignment = centrado
        ws.cell(row=aux, column=2).border = borde
        ws.cell(row=aux, column=3).value = i.proyecto.nombre
        ws.cell(row=aux, column=3).border = borde
        ws.cell(row=aux, column=4).value = i.fecha_ingreso
        ws.cell(row=aux, column=4).alignment = centrado
        ws.cell(row=aux, column=4).border = borde
        ws.cell(row=aux, column=5).value = i.proyecto.desarrolladora.nombre
        ws.cell(row=aux, column=5).border = borde
        ws.cell(row=aux, column=5).alignment = centrado
        ws.cell(row=aux, column=6).value = i.proyecto.desarrolladora.representante
        ws.cell(row=aux, column=6).alignment = centrado
        ws.cell(row=aux, column=6).border = borde
        ws.cell(row=aux, column=7).value = i.proyecto.revisador.first_name + \
            " " + i.proyecto.revisador.last_name
        ws.cell(row=aux, column=7).alignment = centrado
        ws.cell(row=aux, column=7).border = borde
        ws.cell(row=aux, column=8).border = borde
        aux += 1

    nombre_archivo = "Reporte_diario_ingresos_del_" +\
        hoy.strftime("%d_%B_%Y")+"_.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


@ login_required
def reporte_funcionarios(request):
    matriz = []
    fecha_I = ""
    fecha_F = ""
    if request.method == "POST":
        fecha_I = request.POST['inicio']
        fecha_F = request.POST['fin']
        usuarios = User.objects.all()
        labels = []
        atendidos = []
        pendientes = []
        for usuario in usuarios:
            labels.append(usuario.first_name + " " + usuario.last_name)
            proyectos = Proyectos.objects.filter(revisador__pk=usuario.pk)
            print(proyectos)
            aprobados = 0
            reingresos = 0
            for proyecto in proyectos:
                print(proyecto.ingreso)
                if bool(Ingresos.objects.filter(
                        Q(fecha_ingreso__gt=fecha_I) & Q(fecha_ingreso__lt=fecha_F)).filter(proyecto__pk=proyecto.pk).filter(status__pk=2).filter(ingreso=proyecto.ingreso)):
                    aprobados += 1
                elif bool(Ingresos.objects.filter(
                        Q(fecha_ingreso__gt=fecha_I) & Q(fecha_ingreso__lt=fecha_F)).filter(proyecto__pk=proyecto.pk).filter(ingreso=proyecto.ingreso)):
                    print("HOLA")
                    reingresos += 1
            atendidos.append(aprobados)
            pendientes.append(reingresos)

        x = np.arange(len(labels))  # the label locations
        width = 0.15  # the width of the bars

        fig, ax = plt.subplots()
        rects1 = ax.bar(x - width/2, aprobados, width,
                        label='Aprobados', color="#00205c")
        rects2 = ax.bar(x + width/2, pendientes, width,
                        label='Pendientes', color="#0083e6")

        # Add some text for labels, title and custom x-axis tick labels, etc.
        ax.set_ylabel('Proyectos')
        ax.set_title(
            'Registro de actividades por cada funcionario publico periodo ' + fecha_I + ' - ' + fecha_F)
        ax.set_xticks(x)
        ax.set_xticklabels(labels)
        ax.legend()
        for rect in rects1:
            height = rect.get_height()
            ax.annotate('{}'.format(height),
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, 3),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='bottom')

        for rect in rects2:
            height = rect.get_height()
            ax.annotate('{}'.format(height),
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, 3),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='bottom')
        name = "graficas/reporte_funcionario_periodo_del_"+fecha_I+"_al_"+fecha_F+".png"
        fig.tight_layout()
        fig = matplotlib.pyplot.gcf()
        fig.set_size_inches(18.5, 10.5)
        plt.savefig(name)

    return render(request, 'reporte_funcionarios.html', {"datos": matriz, "fechaI": fecha_I, "fechaF": fecha_F})
